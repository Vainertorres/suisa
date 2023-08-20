from django.shortcuts import render
from django.views import generic
from django.db.models import Q
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from datetime import timedelta
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from datetime import date
from django.db.models import Avg, Max, Min, Sum, F, Count

from .models import SegFichaIec, Fichaiec, Conglomerado, ConfigConglomerado

class RepConglomerado(generic.TemplateView):
	template_name='cvd/repconfigconglomerado_form.html'

	def get_context_data(self, **kwargs):
		pk = self.kwargs.get('pk') # El mismo nombre que en tu URL
		context = super().get_context_data(**kwargs)
		cfcong = ConfigConglomerado.objects.filter(pk=pk).first()
		context['panel']='Panel de Administrador'	    
		context['conglo'] = cfcong
		context['obj'] = cfcong.fichaiec.all()
		context['aislados'] = ContactosIec.fichaiec.all()		
		#context['obj'] = ConfigConglomerado.objects.filter(pk=pk).values('')
		return context

class ReportConglomerado(generic.TemplateView):
	template_name='cvd/conglomeradoreport_form.html'

	def municipios_pacientes(self, id):
		config = ConfigConglomerado.objects.filter(pk=id).values('fichaiec','fichaiec__paciente__municipio__descripcion')
		municipio = ""
		for x in config:
			mun = x['fichaiec__paciente__municipio__descripcion']
			if mun != None:
				if municipio == '':
					municipio = mun
				else:
					if mun in municipio:
						pass
					else:
						municipio = municipio + '-'+ mun
		return municipio


	def reporteconglomerado(self):
		pubs = ConfigConglomerado.objects.all()
		diccionario = {}
		data = []
		for p in pubs:		
			municipios=self.municipios_pacientes(p.pk);
			diccionario = {
			'tipoconglomerado':p.conglomerado.tipoconglomerado,
			'direccion':p.conglomerado.direccion,
			'comuna':p.barrio.comuna.descripcion,
			'empresa':p.conglomerado.descripcion,
			'casospos':p.nrocasospos,
			'nrocasosneg':p.nrocasosneg,
			'casospend':p.nrocasospendiente,
			'totalaislados':p.totalperaislados,
			'fuente':p.fuente,
			'municipios':municipios,
			'causacontagio':p.causacontagio,
			'nrocasosrel':p.nrocasosrelacionados, 
			'estado':p.estadoconglomerado,
			'medidasempresa':p.descmedidasinstitucion,
			'medidasecsalud':p.descmeorgasalud,
			'medidascorrecion':p.medidascorrectivas,
			'controlado':p.controlado
			}
			
			data.append(diccionario)
		return data


	def get(self, request,  *args, **kwargs):
		wb = Workbook()
		bandera=True
		if bandera:
			ws=wb.active #Primera hoja activa
			ws.title = "Reporte Conglomerado"
			bandera = False
			ws.merge_cells('A1:R1')	
			ws['A1'] = "INFORME DE CONGLOMERADOS"
			ws['A1'].font = Font(bold=True, size=12)
			ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

			ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['D2'].alignment = Alignment(horizontal="center", vertical="center")

			ws['A2']="NUMERO"
			ws['B2']="CATEGORIA DEL CONGLOMERADO"
			ws['C2']="DIRECCION"
			ws['D2']="COMUNA"
			ws['E2']="NOMBRE DE LA EMPRESA O INSTITUCION"
			ws['F2']="NUMERO DE CASOS POSITIVOS"
			ws['G2']="NUMERO DE CASOS NEGATIVOS"
			ws['H2']="CASOS PENDIENTES"
			ws['I2']="TOTAL DE PERSONAL AISLADO"
			ws['J2']="FUENTE"
			ws['K2']="MUNICIPIOS RELACIONADOS"
			ws['L2']="CAUSA DE CONTAGIO"
			ws['M2']="NUMERO DE CASOS RELACIONADOS"
			ws['N2']="ESTADO DEL CONGLOMERADO"
			ws['O2']="MEDIDAS TOMADAS POR LA EMPRESA"
			ws['P2']="MEDIDAS TOMADAS POR SECRETARIA DE SALUD"
			ws['Q2']="MEDIDAS DE CORRECION"
			ws['R2']="CONTROLADO"

			ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['A2'].font = Font(bold=True, size=12)
			ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['B2'].font = Font(bold=True, size=12)
			ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['C2'].font = Font(bold=True, size=12)
			ws['D2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['D2'].font = Font(bold=True, size=12)
			ws['E2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['E2'].font = Font(bold=True, size=12)
			ws['F2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['F2'].font = Font(bold=True, size=12)
			ws['G2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['G2'].font = Font(bold=True, size=12)
			ws['H2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['H2'].font = Font(bold=True, size=12)
			ws['I2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['I2'].font = Font(bold=True, size=12)
			ws['J2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['J2'].font = Font(bold=True, size=12)
			ws['K2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['K2'].font = Font(bold=True, size=12)
			ws['L2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['L2'].font = Font(bold=True, size=12)
			ws['M2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['M2'].font = Font(bold=True, size=12)
			ws['N2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['N2'].font = Font(bold=True, size=12)
			ws['O2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['O2'].font = Font(bold=True, size=12)
			ws['P2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['P2'].font = Font(bold=True, size=12)
			ws['Q2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['Q2'].font = Font(bold=True, size=12)
			ws['R2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['R2'].font = Font(bold=True, size=12)
			# FIN DEL TITUTLO
			datos = self.reporteconglomerado()
			
			krow = 3
			k = 1
			for cong in datos:
				tipocong = cong['tipoconglomerado']

				ws.cell(row=krow, column=1).value = k				
				ws.cell(row=krow, column=2).value = tipocong.descripcion 
				ws.cell(row=krow, column=3).value = cong['direccion'] 
				ws.cell(row=krow, column=4).value = cong['comuna'] 
				ws.cell(row=krow, column=5).value = cong['empresa']
				ws.cell(row=krow, column=6).value = cong['casospos']
				ws.cell(row=krow, column=7).value = cong['nrocasosneg']
				ws.cell(row=krow, column=8).value = cong['casospend']
				ws.cell(row=krow, column=9).value = cong['totalaislados']
				ws.cell(row=krow, column=10).value = cong['fuente']
				ws.cell(row=krow, column=11).value = cong['municipios']
				ws.cell(row=krow, column=12).value = cong['causacontagio']
				ws.cell(row=krow, column=13).value = cong['nrocasosrel']
				ws.cell(row=krow, column=14).value = cong['estado']
				ws.cell(row=krow, column=15).value = cong['medidasempresa'] 
				ws.cell(row=krow, column=16).value = cong['medidasecsalud']  
				ws.cell(row=krow, column=17).value = cong['medidascorrecion']  
				ws.cell(row=krow, column=18).value = cong['controlado']

				krow += 1
				k += 1

			ws.row_dimensions[1].height = 25
			ws.column_dimensions['O'].width = 60
			ws.column_dimensions['P'].width = 60
			ws.column_dimensions['Q'].width = 60
			ws.column_dimensions['B'].width = 30
			ws.column_dimensions['C'].width = 30
			ws.column_dimensions['E'].width = 30
			ws.column_dimensions['R'].width = 15
			ws.column_dimensions['J'].width = 30
			ws.column_dimensions['K'].width = 30


		#Establecer el nombre del archivo	
		file_name = "conglomerado.xlsx"
		#Establecer el tipo de respuesta que se va a dar
		response = HttpResponse(content_type="application/ms-excel")
		contenido = "attachment; filename={0}".format(file_name)
		response["Content-Disposition"]=contenido
		wb.save(response)
		return response

			
	# def get_context_data(self, **kwargs):
	#     context = super().get_context_data(**kwargs)
	#     context['panel']='Panel de Administrador'
	#     context['obj']= self.reporteconglomerado()
	#     return context



class ReporteCuadroMando(TemplateView):
	def get(self, request, *args, **kwargs):
		nropaccovid = Fichaiec.objects.filter(estado=True).count()
		nropacact = Fichaiec.objects.filter(estado=True, estadoiec="ACT").count()
		nropacrecu = Fichaiec.objects.filter(estado=True, estadoiec="CUR").count()
		nropacfal = Fichaiec.objects.filter(estado=True, estadoiec="FAL").count()

		nropaccasa = 0
		nropachosp = 0
		nropacuci = 0
		nropacspcrpos = Fichaiec.objects.filter(estado=True, respcr="+").count()
		sumdias = Fichaiec.objects.values('paciente').annotate(resta = F('fechaprimuestra') - F('fecinisintomas'))
		sumdiasres = Fichaiec.objects.values('paciente').annotate(resta = F('fechareslab') - F('fechaprimuestra'))
		ambito = Fichaiec.objects.values('ambitoatencion__descripcion').annotate(dcount=Count('ambitoatencion__descripcion')).order_by()


		acumdias = 0
		promdiastomamx = 0
		promdiasres = 0

		kuser = 0
		for totdias in sumdias:
			if totdias['resta']== None:
				pass
			else:
				delta =  totdias['resta']
				if delta.days > 0:
					acumdias += delta.days
				kuser += 1

			if kuser != 0:
				promdiastomamx = acumdias / kuser

		acumdias = 0
		kuser = 0
		for totdias in sumdiasres:
			if totdias['resta'] == None:
				pass
			else:
				delta =  totdias['resta']
				if delta.days > 0:
					acumdias += delta.days
				kuser += 1
			if kuser != 0:
				promdiasres = acumdias / kuser
	
		wb = Workbook()
		bandera=True
		if bandera:
			ws=wb.active #Primera hoja activa
			ws.title = "Cuadro_mando"
			bandera = False
			ws.merge_cells('A1:C1')	
			ws['A1'] = "RESUMEN CUADRO DE MANDO"
			ws['A1'].font = Font(bold=True, size=12)
			ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
									 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
			ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
									top=Side(border_style="thin"), bottom=Side(border_style="thin"))
			ws['C1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
									top=Side(border_style="thin"), bottom=Side(border_style="thin"))
			ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
			ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['C2'].alignment = Alignment(horizontal="center", vertical="center")

			ws['A2']="DESCRIPCIÓN"
			ws['B2']="CANTIDAD"
			ws['C2']=" % "
			ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['A2'].font = Font(bold=True, size=12)
			ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['B2'].font = Font(bold=True, size=12)
			ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
			ws['C2'].font = Font(bold=True, size=12)
			# FIN DEL TITUTLO

			ws['A3'] = "Número de pacientes Diagnosticados con Covid_19:"
			ws['B3'] = nropaccovid			
			ws['A4'] = "PACIENTES SEGÚN ESTADO"
			ws.merge_cells('A4:C4')
			ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
			ws['A4'].font = Font(bold=True, size=12)

			ws['A5'] = "Número de pacientes Activos:"
			ws['B5'] = nropacact			
			ws['A6'] = "Número de pacientes Recuperados:"
			ws['B6'] = nropacrecu
			ws['A7'] = "Número de pacientes Fallecidos:"
			ws['B7'] = nropacfal
			ws['A8'] = "PACIENTES SEGÚN RESULTADO DE LABORATORIO"
			ws.merge_cells('A8:C8')
			ws['A8'].alignment = Alignment(horizontal="center", vertical="center")
			ws['A8'].font = Font(bold=True, size=12)

			ws['A9'] = "Número de pacientes con PCR Positiva:"
			ws.cell(row=9, column=2).value = nropacspcrpos

			ws['A10'] = "PACIENTES SEGÚN AMBITO DE ATENCIÓN"
			ws.merge_cells('A10:C10')
			ws['A10'].alignment = Alignment(horizontal="center", vertical="center")
			ws['A10'].font = Font(bold=True, size=12)

			krow = 10
			for ambate in ambito:				
				krow += 1
				ws.cell(row=krow, column=1).value = ambate['ambitoatencion__descripcion']
				ws.cell(row=krow, column=2).value = ambate['dcount']

			krow += 1
			ws.cell(row=krow, column=1).value = "PROMEDIO DE TIEMPO"
			#ws.merge_cells('A'+str(krow):'C'+str(krow))
			ws.cell(row=krow, column=1).alignment = Alignment(horizontal="center", vertical="center")
			ws.cell(row=krow, column=1).font = Font(bold=True, size=12)
			krow += 1
			ws.cell(row=krow, column=1).value = "Promedio de días desde Inicio de Síntomas Hasta toma de Muestra:"
			ws.cell(row=krow, column=2).value = promdiastomamx

			krow +=1
			ws.cell(row=krow, column=1).value = "Promedio de días desde Toma de Muestra Hasta Resultado:"
			ws.cell(row=krow, column=2).value = promdiasres

			ws.row_dimensions[1].height = 25
			ws.column_dimensions['A'].width = 60
			ws.column_dimensions['B'].width = 12


		#Establecer el nombre del archivo	
		file_name = "cuadrodemando.xlsx"
		#Establecer el tipo de respuesta que se va a dar
		response = HttpResponse(content_type="application/ms-excel")
		contenido = "attachment; filename={0}".format(file_name)
		response["Content-Disposition"]=contenido
		wb.save(response)
		return response

def imprimir_tareas_list(request, f1, f2):
	template_name="cvd/tareasseg_print_all.html"

	fec1=parse_date(f1)
	fec2=parse_date(f2)
	fec2 = fec2 + timedelta(days=1)
	tareas = SegFichaIec.objects.filter(fechaprog__gte=fec1, fechaprog__lt=fec2)
	fec2 = fec2 - timedelta(days=1)
	context={
	'request':request,
	'f1':fec1,
	'f2':fec2,
	'tarea':tareas
	}

	return render(request,template_name,context)
